#--read
#basic method
import pandas as pd

df = pd.read_excel('"C:\Users\shaji\Dropbox\PC\Downloads\household-net-worth-statistics-year-ended-june-2024.csv"')
print(df.head())

#Using openpyxl
import openpyxl

# 1. Setup
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Summary"

# 2. Write Data
# Write headers
sheet['A1'] = "ID"
sheet['B1'] = "Item Name"
sheet['C1'] = "Price"

# Append multiple rows of data
data_rows = [
    (1, "Pen", 1.50),
    (2, "Notebook", 3.25),
    (3, "Eraser", 0.75)
]

for row in data_rows:
    sheet.append(row)

# 3. Save File
workbook.save("inventory_data.xlsx")

print("Excel file 'inventory_data.xlsx' created.")

#Using xlrd
import xlrd

wb = xlrd.open_workbook('file.xls')
sheet = wb.sheet_by_index(0)

# Read a cell
print(sheet.cell_value(0,0))

# Loop through rows
for r in range(sheet.nrows):
    print(sheet.row_values(r))

#Using win32com
import win32com.client as win32

excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(r'C:\path\file.xlsx')
sheet = wb.Sheets('Sheet1')

value = sheet.Cells(1,1).Value  # Row 1, Column 1
print(value)

wb.Close()
excel.Quit()

#writing
#Using pandas
import pandas as pd

data = {'Name': ['Alice', 'Bob', 'Charlie'],
        'Age': [25, 30, 35]}
df = pd.DataFrame(data)

df.to_excel('output.xlsx', index=False)  # index=False avoids writing row numbers

#Using openpyxl
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active
sheet.title = 'MySheet'

# Write headers
sheet['A1'] = 'Name'
sheet['B1'] = 'Age'

# Write rows
sheet.append(['Alice', 25])
sheet.append(['Bob', 30])

wb.save('output.xlsx')

#Using xlsxwriter
import xlsxwriter

workbook = xlsxwriter.Workbook('report.xlsx')
worksheet = workbook.add_worksheet()

# Headers
worksheet.write('A1', 'Name')
worksheet.write('B1', 'Age')

# Data
data = [['Alice', 25], ['Bob', 30], ['Charlie', 35]]
for i, (name, age) in enumerate(data, start=1):
    worksheet.write(i, 0, name)
    worksheet.write(i, 1, age)

workbook.close()

#--visualizing
#Basic line plot
import matplotlib.pyplot as plt

df.plot(x='Name', y='Age', kind='line', marker='o', color='blue')
plt.title('Age of People')
plt.xlabel('Name')
plt.ylabel('Age')
plt.show()

#seaborn
import seaborn as sns
import matplotlib.pyplot as plt

sns.barplot(data=df, x='Name', y='Age', palette='viridis')
plt.title('Age Bar Chart (Seaborn)')
plt.show()

#sns.lineplot() – Line chart
#sns.scatterplot() – Scatter chart
#sns.histplot() – Histogram
#sns.boxplot() – Box plot for statistics
#sns.heatmap() – Correlation heatmap
